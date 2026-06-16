from django.utils import timezone
from django.conf import settings
from datetime import timedelta
from decimal import Decimal
import os
from .utils import percent_change
from .models import Project, MLModel


class InformationClassificationService:
    """Naive Bayes classification services for IMS basic ML workflows."""

    MODEL_PATH = os.path.join(
        settings.BASE_DIR,
        "informationmanagement",
        "ml_models",
        "project_success_naive_bayes.joblib",
    )

    @staticmethod
    def _project_duration_days(project):
        if not project.start_date or not project.end_date:
            return 0
        return max(0, (project.end_date - project.start_date).days)

    @staticmethod
    def _project_label(project):
        """Binary label: 1=likely successful, 0=at risk/unfinished."""
        if project.status == "Completed":
            return 1
        if project.status == "Ongoing" and (project.progress or 0) >= 70:
            return 1
        return 0

    @staticmethod
    def _build_training_rows(projects):
        rows = []
        for project in projects:
            rows.append(
                {
                    "project": project,
                    "category": project.category,
                    "beneficiaries_count": int(project.beneficiaries_count or 0),
                    "progress": int(project.progress or 0),
                    "duration_days": InformationClassificationService._project_duration_days(
                        project
                    ),
                    "label": InformationClassificationService._project_label(project),
                }
            )
        return rows

    @staticmethod
    def _load_model_artifact():
        import joblib

        if not os.path.exists(InformationClassificationService.MODEL_PATH):
            return None
        return joblib.load(InformationClassificationService.MODEL_PATH)

    @staticmethod
    def train_naive_bayes_classifier(lookback_days=3650):
        """Train a basic Naive Bayes classifier for project success prediction."""
        from sklearn.compose import ColumnTransformer
        from sklearn.naive_bayes import GaussianNB
        from sklearn.pipeline import Pipeline
        from sklearn.preprocessing import OneHotEncoder
        from sklearn.model_selection import train_test_split
        from sklearn.metrics import accuracy_score
        import pandas as pd
        import joblib

        since = timezone.now().date() - timedelta(days=lookback_days)
        projects = Project.objects.filter(start_date__gte=since).order_by("id")
        rows = InformationClassificationService._build_training_rows(projects)

        if len(rows) < 8:
            return {
                "trained": False,
                "reason": "Not enough project samples for training",
                "samples": len(rows),
            }

        df = pd.DataFrame(rows)
        X = df[["category", "beneficiaries_count", "progress", "duration_days"]]
        y = df["label"]

        if y.nunique() < 2:
            return {
                "trained": False,
                "reason": "Training labels have only one class",
                "samples": len(rows),
            }

        preprocessor = ColumnTransformer(
            transformers=[
                (
                    "category",
                    OneHotEncoder(handle_unknown="ignore", sparse_output=False),
                    ["category"],
                )
            ],
            remainder="passthrough",
        )

        pipeline = Pipeline(
            steps=[
                ("preprocessor", preprocessor),
                ("classifier", GaussianNB()),
            ]
        )

        accuracy = None
        if len(df) >= 12:
            X_train, X_test, y_train, y_test = train_test_split(
                X,
                y,
                test_size=0.25,
                random_state=42,
                stratify=y,
            )
            pipeline.fit(X_train, y_train)
            y_pred = pipeline.predict(X_test)
            accuracy = round(float(accuracy_score(y_test, y_pred)), 4)
        else:
            pipeline.fit(X, y)

        os.makedirs(
            os.path.dirname(InformationClassificationService.MODEL_PATH), exist_ok=True
        )
        artifact = {
            "model": pipeline,
            "trained_at": timezone.now().isoformat(),
            "feature_names": [
                "category",
                "beneficiaries_count",
                "progress",
                "duration_days",
            ],
        }
        joblib.dump(artifact, InformationClassificationService.MODEL_PATH)

        metric_value = f"Accuracy {accuracy}" if accuracy is not None else "Trained"
        MLModel.objects.update_or_create(
            name="IMS Project Success Classifier",
            defaults={
                "model_type": "Naive Bayes / Classification",
                "status": "Ready",
                "metric": metric_value,
            },
        )

        return {
            "trained": True,
            "samples": len(rows),
            "accuracy": accuracy,
            "model_path": InformationClassificationService.MODEL_PATH,
        }

    @staticmethod
    def predict_project_success(threshold=0.55):
        """Predict success probability and update project prediction fields."""
        import pandas as pd

        artifact = InformationClassificationService._load_model_artifact()
        if not artifact or "model" not in artifact:
            return []

        model = artifact["model"]
        projects = list(Project.objects.all().order_by("-start_date"))
        if not projects:
            return []

        rows = InformationClassificationService._build_training_rows(projects)
        df = pd.DataFrame(rows)
        X = df[["category", "beneficiaries_count", "progress", "duration_days"]]

        probabilities = model.predict_proba(X)[:, 1]
        predictions = []

        for row, probability in zip(rows, probabilities):
            project = row["project"]
            success_percent = round(float(probability) * 100, 2)

            projected_reach = int(
                round(
                    (project.beneficiaries_count or 0)
                    * (0.8 + (float(probability) * 0.4))
                )
            )

            project.predicted_success = Decimal(str(success_percent))
            project.predicted_reach = max(
                projected_reach, project.beneficiaries_count or 0
            )
            project.save(update_fields=["predicted_success", "predicted_reach"])

            if float(probability) >= threshold:
                predictions.append(
                    {
                        "project": project,
                        "confidence": round(float(probability), 4),
                        "predicted_reach": project.predicted_reach,
                        "predicted_label": "Likely Successful",
                    }
                )

        return sorted(predictions, key=lambda item: item["confidence"], reverse=True)

    @staticmethod
    def _fallback_predictions(limit=5):
        projects = Project.objects.all().order_by("-progress", "-beneficiaries_count")[
            :limit
        ]
        fallback = []
        for project in projects:
            heuristic_confidence = min(0.95, max(0.4, (project.progress or 0) / 100))
            fallback.append(
                {
                    "project": project,
                    "confidence": round(float(heuristic_confidence), 4),
                    "predicted_reach": project.beneficiaries_count or 0,
                    "predicted_label": "Heuristic Estimate",
                }
            )
        return fallback

    @staticmethod
    def get_dashboard_snapshot(limit=5):
        """Get prediction rows for dashboard without forcing retrain."""
        model_ready = (
            InformationClassificationService._load_model_artifact() is not None
        )
        predictions = []
        uses_fallback = False

        if model_ready:
            predictions = InformationClassificationService.predict_project_success()

        if not predictions:
            predictions = InformationClassificationService._fallback_predictions(
                limit=limit
            )
            uses_fallback = True

        return {
            "model_ready": model_ready,
            "uses_fallback": uses_fallback,
            "predictions": predictions[:limit],
        }


def get_information_dashboard_data():
    today = timezone.now().date()
    stats = [
        {
            "label": "Active Projects",
            "value": 24,
            "delta": percent_change(24, 21),
            "trend": "up",
        },
        {
            "label": "Partner Orgs",
            "value": 18,
            "delta": percent_change(18, 15),
            "trend": "up",
        },
        {
            "label": "Beneficiaries",
            "value": 1260,
            "delta": percent_change(1260, 1210),
            "trend": "up",
        },
        {
            "label": "Activities This Month",
            "value": 12,
            "delta": percent_change(12, 14),
            "trend": "down",
        },
    ]

    projects = [
        {
            "name": "Community Literacy Boost",
            "category": "Education",
            "status": "Ongoing",
            "start": "Jan 08, 2026",
            "end": "Apr 21, 2026",
            "progress": 72,
        },
        {
            "name": "Agri-Tech Starter Kit",
            "category": "Livelihood",
            "status": "Ongoing",
            "start": "Feb 01, 2026",
            "end": "Jun 30, 2026",
            "progress": 38,
        },
        {
            "name": "Safe Water Initiative",
            "category": "Health",
            "status": "Proposed",
            "start": "Mar 12, 2026",
            "end": "Aug 18, 2026",
            "progress": 15,
        },
    ]

    recent_activities = [
        {
            "title": "Barangay mapping workshop",
            "location": "Brgy. San Miguel",
            "date": today.strftime("%b %d, %Y"),
            "participants": 64,
            "status": "Completed",
        },
        {
            "title": "Partner alignment meeting",
            "location": "Linkages Office",
            "date": today.strftime("%b %d, %Y"),
            "participants": 14,
            "status": "Completed",
        },
        {
            "title": "Baseline survey kick-off",
            "location": "Brgy. Poblacion",
            "date": today.strftime("%b %d, %Y"),
            "participants": 32,
            "status": "Scheduled",
        },
    ]

    sector_coverage = [
        {"name": "Education", "key": "education", "value": 32},
        {"name": "Health", "key": "health", "value": 24},
        {"name": "Livelihood", "key": "livelihood", "value": 18},
        {"name": "Environment", "key": "environment", "value": 14},
        {"name": "Governance", "key": "governance", "value": 12},
    ]

    return {
        "stats": stats,
        "projects": projects,
        "recent_activities": recent_activities,
        "sector_coverage": sector_coverage,
    }


def get_projects_data():
    return {
        "filters": ["Education", "Health", "Livelihood", "Environment", "Governance"],
        "projects": [
            {
                "name": "Community Literacy Boost",
                "category": "Education",
                "lead": "Prof. C. Reyes",
                "status": "Ongoing",
                "beneficiaries": 420,
                "timeline": "Jan - Apr 2026",
            },
            {
                "name": "Agri-Tech Starter Kit",
                "category": "Livelihood",
                "lead": "Dr. L. Santos",
                "status": "Ongoing",
                "beneficiaries": 210,
                "timeline": "Feb - Jun 2026",
            },
            {
                "name": "Safe Water Initiative",
                "category": "Health",
                "lead": "Dr. M. Dela Cruz",
                "status": "Proposed",
                "beneficiaries": 130,
                "timeline": "Mar - Aug 2026",
            },
            {
                "name": "Eco-Youth Engagement",
                "category": "Environment",
                "lead": "Ms. P. Valdez",
                "status": "Completed",
                "beneficiaries": 180,
                "timeline": "Sep - Dec 2025",
            },
        ],
    }


def get_beneficiaries_data():
    return {
        "segments": [
            {"label": "Women-led households", "value": 34},
            {"label": "Senior citizens", "value": 18},
            {"label": "Youth", "value": 26},
            {"label": "PWD", "value": 7},
            {"label": "Farmer groups", "value": 15},
        ],
        "communities": [
            {"name": "Brgy. San Miguel", "households": 220, "priority": "High"},
            {"name": "Brgy. Poblacion", "households": 180, "priority": "Medium"},
            {"name": "Brgy. Mabini", "households": 160, "priority": "Medium"},
            {"name": "Brgy. Bagong Silang", "households": 140, "priority": "Low"},
        ],
    }


def get_partners_data():
    return {
        "partners": [
            {
                "name": "Local Health Unit",
                "type": "Government",
                "status": "Active",
                "engagement": "High",
                "contribution": "Medical missions",
            },
            {
                "name": "AgriCoop Federation",
                "type": "NGO",
                "status": "Active",
                "engagement": "Medium",
                "contribution": "Inputs and training",
            },
            {
                "name": "Green Future Foundation",
                "type": "Private",
                "status": "Prospecting",
                "engagement": "Low",
                "contribution": "CSR funding",
            },
        ],
        "pipeline": [
            {"stage": "Prospecting", "count": 6},
            {"stage": "Negotiation", "count": 3},
            {"stage": "Active", "count": 18},
            {"stage": "Dormant", "count": 4},
        ],
    }


def get_activities_data():
    return {
        "upcoming": [
            {
                "title": "Data collection training",
                "date": "Feb 12, 2026",
                "location": "CES Hub",
                "owner": "Extension Unit",
                "status": "Scheduled",
            },
            {
                "title": "Partner alignment meeting",
                "date": "Feb 15, 2026",
                "location": "Board Room",
                "owner": "Linkages Unit",
                "status": "Scheduled",
            },
        ],
        "recent": [
            {
                "title": "Baseline survey kickoff",
                "date": "Feb 02, 2026",
                "location": "Brgy. Poblacion",
                "owner": "Research Unit",
                "status": "Completed",
            },
            {
                "title": "Impact narrative workshop",
                "date": "Jan 28, 2026",
                "location": "CES Hub",
                "owner": "QA Team",
                "status": "Completed",
            },
        ],
    }


def get_analytics_data():
    return {
        "kpis": [
            {"label": "Program Reach", "value": "1,260", "trend": "+4.1%"},
            {"label": "Partner Engagement", "value": "76%", "trend": "+2.2%"},
            {"label": "Completion Rate", "value": "68%", "trend": "+5.8%"},
            {"label": "Satisfaction", "value": "4.6/5", "trend": "+0.3"},
        ],
        "insights": [
            "Education projects have the highest retention rate in the last two quarters.",
            "Health initiatives show rising demand in coastal barangays.",
            "Livelihood programs benefit most when paired with partner micro-finance support.",
        ],
        "trend": [
            {"label": "Q1", "value": 62},
            {"label": "Q2", "value": 68},
            {"label": "Q3", "value": 71},
            {"label": "Q4", "value": 75},
        ],
    }


def get_reports_data():
    return {
        "reports": [
            {
                "title": "Accomplishment Report",
                "period": "Q4 2025",
                "status": "Generated",
                "owner": "QA Team",
            },
            {
                "title": "Impact Evaluation",
                "period": "CY 2025",
                "status": "In Review",
                "owner": "Research Unit",
            },
            {
                "title": "Partner Contribution",
                "period": "Jan 2026",
                "status": "Draft",
                "owner": "Linkages Unit",
            },
        ],
        "templates": [
            "Extension Activity Report",
            "Partner Commitment Tracker",
            "Beneficiary Outcome Matrix",
            "Accreditation Evidence Bundle",
        ],
    }


def get_ml_lab_data():
    return {
        "models": [
            {
                "name": "Program Reach Predictor",
                "type": "Regression",
                "status": "Prototype",
                "metric": "MAE 0.18",
            },
            {
                "name": "Beneficiary Segmentation",
                "type": "Clustering",
                "status": "Ready",
                "metric": "Silhouette 0.62",
            },
            {
                "name": "Partner Engagement Classifier",
                "type": "Classification",
                "status": "Training",
                "metric": "F1 0.74",
            },
        ],
        "pipelines": [
            {"name": "Data Cleaning", "status": "Operational"},
            {"name": "Feature Store", "status": "Design"},
            {"name": "Model Registry", "status": "Backlog"},
            {"name": "Monitoring", "status": "Planned"},
        ],
        "experiments": [
            {
                "name": "Impact Score v2",
                "owner": "ML Team",
                "status": "Running",
                "updated": "2 hours ago",
            },
            {
                "name": "Engagement uplift test",
                "owner": "Data Science",
                "status": "Completed",
                "updated": "Yesterday",
            },
        ],
    }


# ============================================================================
# Feature 2: Financial Reporting and Export Services
# ============================================================================


class FinancialReportingService:
    """Service for generating comprehensive financial reports and analytics"""

    @staticmethod
    def get_association_financial_summary(date_from=None, date_to=None):
        """Generate association-wide financial summary"""
        from .models import FundAllocation, FundExpense, ContributionFund
        from django.db.models import Sum
        
        query_allocations = FundAllocation.objects.all()
        query_expenses = FundExpense.objects.all()
        
        if date_from:
            query_allocations = query_allocations.filter(allocated_date__gte=date_from)
            query_expenses = query_expenses.filter(expense_date__gte=date_from)
        
        if date_to:
            query_allocations = query_allocations.filter(allocated_date__lte=date_to)
            query_expenses = query_expenses.filter(expense_date__lte=date_to)
        
        total_allocated = query_allocations.aggregate(
            total=Sum("amount")
        )["total"] or Decimal("0.00")
        
        total_expenses = query_expenses.aggregate(
            total=Sum("amount")
        )["total"] or Decimal("0.00")
        
        available_funds = total_allocated - total_expenses
        
        return {
            "total_allocations": total_allocated,
            "total_expenses": total_expenses,
            "available_funds": available_funds,
            "num_active_funds": ContributionFund.objects.filter(status="active").count(),
            "date_from": date_from,
            "date_to": date_to,
        }

    @staticmethod
    def get_fund_utilization_report(date_from=None, date_to=None):
        """Generate detailed fund utilization report per project"""
        from .models import ContributionFund
        
        funds = ContributionFund.objects.all()
        
        report_data = []
        for fund in funds:
            allocated = fund.get_total_allocated()
            used = fund.get_total_used()
            remaining = fund.get_remaining_balance()
            utilization_rate = (used / allocated * 100) if allocated > 0 else 0
            
            report_data.append({
                "fund_name": fund.name,
                "fund_id": fund.id,
                "budget_required": fund.budget_required,
                "total_allocated": allocated,
                "total_used": used,
                "remaining_balance": remaining,
                "utilization_rate": round(utilization_rate, 2),
                "status": fund.status,
            })
        
        return report_data

    @staticmethod
    def get_monthly_contribution_report(year, month):
        """Generate monthly contribution report"""
        from .models import FundAllocation
        from dateutil import relativedelta
        from datetime import date
        
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        allocations = FundAllocation.objects.filter(
            allocated_date__range=[start_date, end_date]
        )
        
        total = allocations.aggregate(Sum("amount"))["amount__sum"] or Decimal("0.00")
        count = allocations.count()
        
        return {
            "year": year,
            "month": month,
            "start_date": start_date,
            "end_date": end_date,
            "total_contributions": total,
            "number_of_allocations": count,
            "allocations": allocations,
        }

    @staticmethod
    def get_annual_financial_report(year):
        """Generate annual financial report"""
        from .models import FundAllocation, FundExpense
        from datetime import date
        
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        
        total_allocated = FundAllocation.objects.filter(
            allocated_date__range=[start_date, end_date]
        ).aggregate(Sum("amount"))["amount__sum"] or Decimal("0.00")
        
        total_expenses = FundExpense.objects.filter(
            expense_date__range=[start_date, end_date]
        ).aggregate(Sum("amount"))["amount__sum"] or Decimal("0.00")
        
        monthly_data = []
        for month in range(1, 13):
            month_allocations = FundAllocation.objects.filter(
                allocated_date__year=year,
                allocated_date__month=month
            ).aggregate(Sum("amount"))["amount__sum"] or Decimal("0.00")
            
            month_expenses = FundExpense.objects.filter(
                expense_date__year=year,
                expense_date__month=month
            ).aggregate(Sum("amount"))["amount__sum"] or Decimal("0.00")
            
            monthly_data.append({
                "month": month,
                "allocations": month_allocations,
                "expenses": month_expenses,
            })
        
        return {
            "year": year,
            "total_allocated": total_allocated,
            "total_expenses": total_expenses,
            "available_funds": total_allocated - total_expenses,
            "monthly_breakdown": monthly_data,
        }


class ExportService:
    """Service for exporting reports to various formats"""

    @staticmethod
    def export_to_excel(report_data, report_type, filename):
        """Export report data to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return None
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type
        
        # Add header
        ws['A1'] = f"{report_type} Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        # Add data
        if isinstance(report_data, list):
            headers = list(report_data[0].keys()) if report_data else []
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=3, column=col_idx, value=header).font = Font(bold=True)
            
            for row_idx, row_data in enumerate(report_data, 4):
                for col_idx, value in enumerate(row_data.values(), 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(filename)
        return filename

    @staticmethod
    def export_to_pdf(report_data, report_type, filename):
        """Export report data to PDF"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
        except ImportError:
            return None
        
        doc = SimpleDocTemplate(filename, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph(f"<b>{report_type} Report</b>", styles['Heading1'])
        story.append(title)
        story.append(Spacer(1, 0.3 * inch))
        
        # Create table data
        if isinstance(report_data, list) and report_data:
            headers = [[str(h) for h in report_data[0].keys()]]
            data = headers + [[str(v) for v in row.values()] for row in report_data]
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(table)
        
        doc.build(story)
        return filename


class MemberContributionService:
    """Service for member contribution tracking and monitoring"""

    @staticmethod
    def calculate_payment_status(member_record):
        """Calculate and update payment status based on due amount"""
        from datetime import date, timedelta
        
        today = date.today()
        
        if member_record.due_amount == 0:
            return "on_time"
        
        if member_record.last_payment_date:
            days_overdue = (today - member_record.last_payment_date).days
            if days_overdue > 90:
                return "delinquent"
            elif days_overdue > 30:
                return "overdue"
        
        return "on_time"

    @staticmethod
    def generate_member_statement(member_record):
        """Generate individual member contribution statement"""
        from .models import MemberContributionRecord
        
        payment_status = MemberContributionService.calculate_payment_status(member_record)
        member_record.payment_status = payment_status
        member_record.save()
        
        return {
            "member_name": member_record.member_name,
            "employee_id": member_record.employee_id,
            "department": member_record.department,
            "total_contributions": member_record.total_contributions,
            "current_balance": member_record.current_balance,
            "due_amount": member_record.due_amount,
            "late_payment_penalties": member_record.late_payment_penalties,
            "payment_status": payment_status,
            "last_payment_date": member_record.last_payment_date,
            "generated_at": timezone.now(),
        }

    @staticmethod
    def get_filtered_members(filters):
        """Get members filtered by specified criteria"""
        from .models import MemberContributionRecord
        
        query = MemberContributionRecord.objects.all()
        
        if filters.get('status_filter') and filters['status_filter'] != 'all':
            query = query.filter(payment_status=filters['status_filter'])
        
        if filters.get('member_name'):
            query = query.filter(member_name__icontains=filters['member_name'])
        
        if filters.get('employee_id'):
            query = query.filter(employee_id__icontains=filters['employee_id'])
        
        if filters.get('department'):
            query = query.filter(department=filters['department'])
        
        if filters.get('date_from'):
            query = query.filter(created_at__gte=filters['date_from'])
        
        if filters.get('date_to'):
            query = query.filter(created_at__lte=filters['date_to'])
        
        return query.order_by('-updated_at')

    @staticmethod
    def export_member_statement(member_record, format='pdf'):
        """Export individual member statement"""
        statement = MemberContributionService.generate_member_statement(member_record)
        
        filename = f"member_statement_{member_record.employee_id}_{timezone.now().strftime('%Y%m%d')}"
        
        if format == 'excel':
            filename += '.xlsx'
            ExportService.export_to_excel([statement], 'Member Statement', filename)
        elif format == 'pdf':
            filename += '.pdf'
            ExportService.export_to_pdf([statement], 'Member Statement', filename)
        
        return filename
